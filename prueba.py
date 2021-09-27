# Nombre archivo mercado pago: Informe Mercado Pago.xlsx
# Deposits Summary: Nombre de la columna del archivo de excel de eventtia donde aparece el "EXTERNEL_REFERENCE" del archivo de mercado pago.


### DEJAR LOS SIGUIENTES NOMBRES A LAS COLUMNAS ###
# CC: Columna de las cedulas.
# NIT: Columna para cedula o nit para facturación.
# NOM: Columna facturar a nombre de.
# DIR: Columna de dirección para facturación.
### CAMBIAR EL NOMBRE DE LA PESTAÑA DE EXCEL POR: Sheet1


import openpyxl
from openpyxl.styles import Font
import os
import pandas as pd
import re


# Formatear el telefono para quetarle el número de país.
def telephone_format(telephone):
    telephoneNumRegex = re.compile(r'(\d\d)(\d{10})')
    mo = telephoneNumRegex.search(str(telephone))
    if mo.group(1) == "57":
        return mo.group(2)
    else:
        return telephone

def email_format(email):
    emailNumRegex = re.compile(r'[\w\._]{2,30}\+?[\w]{0,10}@[\w\.\-]{2,}\.\w{2,6}')
    if re.match(emailNumRegex, email):
        return "CORRECTO"
    else:
        return "INCORRECTO"

def dir_format(dir):
    crNumRegex = re.compile(r'[cC]+[a-zA-Z]*[rR]+[a-zA-Z]*\D')
    clNumRegex = re.compile(r'[cC]+[a-zA-Z]*[lL]+[a-zA-Z]*\D')
    tvNumRegex = re.compile(r'[tT]+[a-zA-Z]*[vV]+[a-zA-Z]*\D')
    avNumRegex = re.compile(r'[aA]+[a-zA-Z]*[vV]+[a-zA-Z]*\D')
    dgNumRegex = re.compile(r'[dD]+[a-zA-Z]*[gG]+[a-zA-Z]*\D')
    dirSplit = dir.split()
    for i in range(len(dirSplit)):
        if re.match(crNumRegex, dirSplit[i]):
            dirSplit[i] = "CR"
        elif re.match(clNumRegex, dirSplit[i]):
            dirSplit[i] = "CL"
        elif re.match(tvNumRegex, dirSplit[i]):
            dirSplit[i] = "TV"
        elif re.match(avNumRegex, dirSplit[i]):
            dirSplit[i] = "AV"
        elif re.match(dgNumRegex, dirSplit[i]):
            dirSplit[i] = "DG"
    return ' '.join(dirSplit)

def nit_format(nit):
    nitNumRegex = re.compile(r'.*(\d{6,})-\d')
    if re.match(nitNumRegex, str(nit)):
        mo = nitNumRegex.search(str(nit))
        return mo.group(1)
    else:
        return nit

    

# archivo con información de mercado pago.
mercado_pago_drive = "Informe Mercado Pago.xlsx"

# archivo con la información de registros de eventia.
eventtia_drive = "eventtia.xlsx"

# archivo con información predefinida del evento.
modificaciones_drive = "modificaciones.xlsx"

# archivo para facturación electrónica.
resultados_drive = openpyxl.load_workbook("resultados.xlsx")
resultados_sheet = resultados_drive['Sheet1']

mercado_libre_sheet = pd.read_excel(mercado_pago_drive, sheet_name="Sheet0")
modificaciones_sheet = pd.read_excel(modificaciones_drive, sheet_name="Sheet1")
eventtia_sheet = pd.read_excel(eventtia_drive, sheet_name="Sheet1")


# Nombre Columnas
nombre_columnas = ["FE", "RC", "CD", "RAZON SOCIAL/NOMBRES Y APELLIDOS", "NIT", "DIV", "FECHA REGISTRO", "CIUDAD", "DOMICILIO PRINCIPAL", "CONTACTO DE FACTURACIÓN", "TELEFONO", "E-MAIL DE FACTURACIÓN", "CENTRO DE COSTOS", "PRODUCTO SIIGO", "DESCRIPCIÓN PARA FACTURACIÓN", "FE", "SUBTOTAL", "IVA", "VALOR TOTAL", "No. FACTURAS", "RESPONSABLE ENDEAVOR", "", "", "","NIT ORIGINAL", "DOMICILIO ORIGINAL", "TELEFONO ORIGINAL", "POSIBLES ERROR EMAIL"]




# Recopilar "EXTERNEL_REFERENCE" del reporte de mercado pago en un array.
externalReference = []

for i in range(len(mercado_libre_sheet)):
    if mercado_libre_sheet['EXTERNAL_REFERENCE'][i] not in externalReference:
        externalReference.append((mercado_libre_sheet['EXTERNAL_REFERENCE'][i]))


# Recopilar información del archivo "modificaciones.xlsx"
modificaciones = []
for i in range(len(modificaciones_sheet)):
    modificaciones.append(modificaciones_sheet['MODIF'][i])



# Guardar información del archivo de evenntia con referencia al EXTERNAL_REFERENCE del archivo de mercado pago.
datos = []

for i in range(len(externalReference)):
    for j in range(len(eventtia_sheet)):
        if len(str(eventtia_sheet['Deposits Summary'][j]).split()) >= 4:
            if externalReference[i] == str(eventtia_sheet['Deposits Summary'][j]).split()[4]:
                datos.append([eventtia_sheet['First Name'][j], eventtia_sheet['Last Name'][j], eventtia_sheet['CC'][j], eventtia_sheet['Email'][j], eventtia_sheet['Telephone'][j] ,eventtia_sheet['City'][j], eventtia_sheet['NOM'][j], eventtia_sheet['NIT'][j], eventtia_sheet['DIR'][j], eventtia_sheet['Total'][j]])


# Titulos de las columnas
for i in range(len(nombre_columnas)):
    resultados_sheet.cell(row=1, column=i+1).value = nombre_columnas[i]
    # Formato bold al título de las columnas.
    resultados_sheet.cell(row=1, column=i+1).font = Font(bold=True)


# Ingresar los datos al archivo resultados.xlsx
for i in range(len(datos)):
    resultados_sheet.cell(row=i+2, column=4).value = str.upper(datos[i][6])
    resultados_sheet.cell(row=i+2, column=5).value = str(nit_format(datos[i][7]))
    resultados_sheet.cell(row=i+2, column=7).value = modificaciones[0]
    resultados_sheet.cell(row=i+2, column=8).value = str.upper(datos[i][5])
    resultados_sheet.cell(row=i+2, column=9).value = str.upper(dir_format(datos[i][8]))
    resultados_sheet.cell(row=i+2, column=10).value = str.upper(datos[i][0]) + " " + str.upper(datos[i][1])
    resultados_sheet.cell(row=i+2, column=11).value = str(telephone_format(datos[i][4]))
    resultados_sheet.cell(row=i+2, column=12).value = str(datos[i][3])
    resultados_sheet.cell(row=i+2, column=13).value = modificaciones[1]
    resultados_sheet.cell(row=i+2, column=14).value = modificaciones[2]
    resultados_sheet.cell(row=i+2, column=15).value = modificaciones[3]
    resultados_sheet.cell(row=i+2, column=17).value = round(datos[i][9]*0.84034)
    resultados_sheet.cell(row=i+2, column=18).value = round(datos[i][9]-(datos[i][9]*0.84034))
    resultados_sheet.cell(row=i+2, column=19).value = datos[i][9]
    resultados_sheet.cell(row=i+2, column=20).value = modificaciones[4]
    resultados_sheet.cell(row=i+2, column=21).value = modificaciones[5]
    ### Valores Originales
    resultados_sheet.cell(row=i+2, column=25).value = str(datos[i][7])
    resultados_sheet.cell(row=i+2, column=26).value = str.upper(datos[i][8])
    resultados_sheet.cell(row=i+2, column=27).value = str(datos[i][4])
    resultados_sheet.cell(row=i+2, column=28).value = str(email_format(datos[i][3]))
    


resultados_drive.save("resultados.xlsx")